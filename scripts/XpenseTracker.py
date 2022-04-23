""" This script should uphold the following format and purpose:

1. Ask the user what they they are doing with their expenses tracking
    - Adding expenses to an existing sheet
    - Creating a new expenses sheet
    - Deleting a current expenses sheet
    - Examining expenses across multiple (or a single) sheet(s) (should just be a printout of expenses, totals for each month, and combined expenses for each month)
        - In addition, we can optionally show a percentage breakdown for each category of expenses
2. Ask the user for input as to which excel spreadsheet they are using to input expense information (give a list of current sheets)
    - if the user inputs a sheet that isn't there, a new sheet is created
    - once the user inputs an excel sheet, we run a loop until the user says that they're finished inputting items. There will be separate loops for costs and expenses to allow continuity and quick input
    - give an exit input at all stages of the for loop (memo, date, amount) such as -exit
3. All arguments should be able to be inputted via command line
    - memo arguments for costs should be put in the first empty space of column A
    - date arguments for costs should be put in the first empty space of column B under B3
    - Amount arguments for costs should be put in the first empty space of column C under C3
    - Same for Income for columns D, E and F, respectively
    - After inputting each costs/expense, the program should display "Income", "Money spent", and "Balance (Income-money spent)"
"""
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle
from openpyxl import Workbook
import easygui
import sys
import os

date_style = NamedStyle(name='datetime', number_format='YYYY-MM-DD')
# Creating main() function as framework for making code more readable

def main():

    While True:
        # User input determines what function is called depending on the task being done
        action = easygui.buttonbox("Welcome to Xpense tracker! What would you like to do today?", 'Action', 
        ('Add Expense/Income', 'Remove Expense/Income', 'Make New Expense Sheet', 'Delete Expense Sheet', 'Compare Expense Sheets', 'Create New Expense Workbook'))

        if (action == 'Add Expense/Income'):
            AddExpenseOrIncome() # TODO: this function needs to be built

        elif (action == 'Remove Expense/Income'):
            RemoveExpenseOrIncome() # TODO: this function needs to be built
        
        elif (action == 'Make New Expense Sheet'):
            MakeExpenseSheet() # TODO: this function needs to be built

        elif (action == 'Delete Expense Sheet'):
            DeleteExpenseSheet() # TODO: this function needs to be built

        elif (action == 'Compare Expense Sheets'):
            CompareExpenseSheets() # TODO: this function needs to be built
        
        elif (action == 'Create New Expense Workbook'):
            CreateNewExpenseWorkbook() # TODO: this function needs to be built

        else:
            print("Please enter a valid response from the options in the dialog box") # TODO: this function needs to be built
    
    action = easygui.buttonbox("Would you like to continue with tracking you expenses?", "Continue?", ('yes', 'no'))

    if (action == 'yes'):
        continue
    if (action == 'no'):
        print("Thank you for using XpenseTracker!")
        sys.exit()
    


    

    
def CreateNewExpenseWorkbook():

    print("Your expenses spreadsheet will be saved in the same directory where you run XpenseTracker.py")
    print("What would you like to call your expenses workbook?")
    wbName = input(">   ")
    print("Here is the folder where you can find you new expenses workbook")
    print(os.getcwd())
    wb = Workbook()
    wbFile = wbName + '.xlsx'

    MakeExpenseSheet(wb)

    if ('Sheet' in wb.sheetnames):
        del wb['Sheet']

    wb.save(wbFile)


def MakeExpenseSheet(xlsxWorkbook):
    
    newSheet = True
    while newSheet:
        print("What would you like to call your new expense sheet?")
        sheetName = input(">    ")
        xlsxWorkbook.create_sheet(sheetName)

        # Set up formatting of your sheet
        ws = xlsxWorkbook[sheetName]
        ws['A1'] = sheetName
        ws.merge_cells('A1:F1')
        ws['A2'] = 'Costs'
        ws['D2'] = 'Income'
        ws.merge_cells('A2:C2')
        ws.merge_cells('D2:F2')
        ws['A3'] = 'Memo'
        ws['B3'] = 'Date'
        ws['C3'] = 'Amount'
        ws['D3'] = 'Memo'
        ws['E3'] = 'Date'
        ws['F3'] = 'Amount'
        ws['H3'] = 'Balance'
        ws['I3'] = 'Spent'
        ws['J3'] = 'Income'

        print("would you like to make any other new sheets in your workbook? 'y' or 'n'")
        anotherSheet = input(">    ").lower()


        ValidInput = False
        while ValidInput == False:
            if (anotherSheet == "y"):
                ValidInput = True
                continue
            elif (anotherSheet == "n"):
                ValidInput = True
                newSheet = False
            else:
                print("Invalid response. Please enter 'y' to make another sheet or 'n' to stop making sheets.")




    

    
