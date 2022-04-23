""" This script should uphold the following format and purpose:

1. Ask the user what they they are doing with their expenses tracking
    - Adding expenses to an existing sheet
    - Creating a new expenses sheet
    - Deleting a current expenses sheet
    - Examining expenses across multiple (or a single) sheet(s) (should just be a printout of expenses, totals for each month, and combined expenses for each month)
        - In addition, we can optionally show a percentage breakdown for each category of expenses
2. Ask the user for input as to which excel spreadsheet they are using to input expense information (give a list of current sheets)
    - if the user inputs a sheet that isn't there, a new sheet is created
    - once the user inputs an excel sheet, we run a loop until the user says that they're finished inputting items. There will be separate loops for costs and expenses to allow continuity and quick input
    - give an exit input at all stages of the for loop (memo, date, amount) such as -exit
3. All arguments should be able to be inputted via command line
    - memo arguments for costs should be put in the first empty space of column A
    - date arguments for costs should be put in the first empty space of column B under B3
    - Amount arguments for costs should be put in the first empty space of column C under C3
    - Same for Income for columns D, E and F, respectively
    - After inputting each costs/expense, the program should display "Income", "Money spent", and "Balance (Income-money spent)"
"""
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle
import easygui
import sys

# set up date style
date_style = NamedStyle(name='datetime', number_format='YYYY-MM-DD')

# Xpense tracker
action = easygui.buttonbox("""Welcome to Xpense tracker! What would you like to do today?""", 'Action', 
('Add Expense', 'Remove expenses', 'Make New Sheet', 'Delete Expense Sheet', 'Compare Multiple Sheets'))

wbPath = "C:\\Users\\keess\\Desktop\\2021_Budget_spreadsheet.xlsx"
# loading workbook to insert new sheets
wb = load_workbook(wbPath,)

if (action == "Make New Sheet"):
    print("Current sheet names: ")
    print(wb.sheetnames)
    sheetName = input("What would you like to name your new sheet?: ")
    target = wb['template']
    wb.copy_worksheet(target)
    wb_sheet = wb["template Copy"]
    wb_sheet.title = sheetName

    MonthNum = input("What is the number corresponding to this month (1-12):    ")
    Year = input("What is the year?:    ")

    askPay = input("The template has your monthly income at 2 paychecks of " + str(target["F4"].value) + " Are you expecting the same amount this month? [y] or [n] ")
    # insert functionality for adding additional sources of income besides salary

    if (askPay == "n"):
        # collect information on pay
        payVal = input("How much money do you expect to take home from each paycheck?   ")
        payVal = float(payVal)
        numChecks = input("How many paychecks are you expecting to receive this month?  ")
        numChecks = int(numChecks)

        for i in range(numChecks):
            Amount_cell = "F" + str(i+4)
            Memo_cell = "D" + str(i+4)
            Date_cell = "E" + str(i+4)
            wb_sheet[Amount_cell] = payVal
            wb_sheet[Amount_cell].number_format = '0.00E+00'
            wb_sheet[Memo_cell] = "Salary"
            wb_sheet[Date_cell] = Year + "-" + MonthNum + "-01"
            wb_sheet[Date_cell].style = date_style

    else:

        for i in range(2):
            Date_cell = "E" + str(i+4)
            wb_sheet[Date_cell] = MonthNum + "/1/" + Year
        

    askRent = input("Are your rent and utilities about the same as the template value of " + str(target["C4"].value) + "? [y] or [n]   ")

    if (askRent == "n"):
        rentVal = input("How much do you expet to pay on combined rent and utilities?   ")
        Amount_cell = "C4"
        Date_cell = "B4"
        wb_sheet[Date_cell] = MonthNum + "/1/" + Year
        wb_sheet[Amount_cell] = float(rentVal)

    else:
        Date_cell = "B4"
        wb_sheet[Date_cell] = MonthNum + "/1/" + Year
    
    wb.save(wbPath) 
    print("Awesome! We've inputted this data into your new sheet called " + sheetName)

if (action == "Add Expense"):
    # ask if the user wants to add expenses or income
    print("Current sheet names: ")
    print(wb.sheetnames)
    sheetName = easygui.buttonbox("What sheet would you like to work in?", "sheets", wb.sheetnames)
    sheet = wb[sheetName]
    inputType = input("Would you like to add expenses or income?", 'type of entry', ('expenses', 'income'))
    inputType = int(inputType)

    # adding expenses
    if (inputType == 'expenses'):
        
        while True:
            
            max_row_expense = max((exp.row for exp in sheet['A'] if exp.value is not None))
            row = str(max_row_expense + 1)

            memo_cell = "A" + row
            date_cell = "B" + row
            amount_cell = "C" + row

            exp_memo = input("Input this expense's memo:   ")
            exp_date = input("Input this expense's date in YYYY-MM-DD format:    ")
            exp_amount = input("Input this expense's amount:    ")

            sheet[memo_cell] = exp_memo
            sheet[date_cell] = exp_date
            sheet[date_cell].style = 'datetime'
            sheet[amount_cell] = float(exp_amount)

            # checking values previously entered for expenses
            for i in range(4, int(row)+1):
                disp_row = i
                memo_cell = "A" + str(disp_row)
                date_cell = "B" + str(disp_row)
                amount_cell = "C" + str(disp_row)

                print(sheet[memo_cell].value + "  " + sheet[date_cell].value + "    " + str(sheet[amount_cell].value))

            stopCondition = input("would you like to keep inputting expenses? [y] to continue or [n] to quit    ")
            if (stopCondition == "n"):
                break
    
    if (inputType == 'income'):
        
        while True:

            max_row_income = max((inc.row for inc in sheet['D'] if inc.value is not None))

            memo_cell = "D" + row
            date_cell = "E" + row
            amount_cell = "C" + row

            inc_memo = input("Input this income's memo:   ")
            inc_date = input("Input this income's date in YYYY-MM-DD format:    ")
            exp_amount = input("Input the income amount:    ")

            sheet[memo_cell] = inc_memo
            sheet[date_cell] = inc_date
            sheet[date_cell].style = 'datetime'
            sheet[amount_cell] = float(inc_amount)

            for i in range(4, int(row)+1):
                disp_row = i
                memo_cell = "D" + str(disp_row)
                date_cell = "E" + str(disp_row)
                amount_cell = "F" + str(disp_row)

                print(sheet[memo_cell].value + "  " + sheet[date_cell].value + "  " + str(sheet[amount_cell].value))

            stopCondition = easygui.ynbox("would you like to keep inputting expenses?", 'Choice', ('yes', 'no'))
            if (stopCondition == "n"):
                break

            

wb.save(wbPath) 
